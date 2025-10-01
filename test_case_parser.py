import re
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import datetime

def parse_test_cases_from_file(file_path):
    """
    Lee un archivo txt y parsea los casos de prueba
    Detecta automáticamente si es formato TAB-delimited o texto corrido
    """
    try:
        with open(file_path, 'r', encoding='utf-8') as file:
            content = file.read()
        
        lines = content.strip().split('\n')
        if not lines:
            return []
        
        # Detectar si es formato TAB-delimited
        first_line = lines[0]
        if '\t' in first_line and ('Nombre' in first_line or 'Secuencia' in first_line):
            print("Detectado formato TAB-delimited (CSV/Excel)")
            return parse_tab_delimited(lines)
        else:
            print("Detectado formato texto corrido")
            return parse_text_format(content)
    
    except Exception as e:
        print(f"Error leyendo el archivo: {e}")
        return []

def parse_tab_delimited(lines):
    """
    Parsea formato separado por tabs (como el que muestras en los ejemplos)
    """
    parsed_cases = []
    
    # Buscar la línea de headers
    header_line_idx = -1
    for i, line in enumerate(lines):
        if '\t' in line and any(keyword in line for keyword in ['Nombre', 'Secuencia', 'Datos', 'Resultado']):
            header_line_idx = i
            break
    
    if header_line_idx == -1:
        print("No se encontraron headers válidos")
        return []
    
    # Extraer headers
    headers = lines[header_line_idx].split('\t')
    print(f"Headers detectados: {headers}")
    
    # Procesar cada línea de datos
    for i in range(header_line_idx + 1, len(lines)):
        line = lines[i].strip()
        if not line:
            continue
        
        columns = line.split('\t')
        
        # Asegurar que tenemos al menos 4 columnas
        while len(columns) < 4:
            columns.append("")
        
        case_data = {
            'Nombre': columns[0].strip() if len(columns) > 0 else "",
            'Secuencia': columns[1].strip() if len(columns) > 1 else "",
            'Datos requeridos / Datos Adicionales': columns[2].strip() if len(columns) > 2 else "",
            'Resultado Esperado': columns[3].strip() if len(columns) > 3 else ""
        }
        
        # Solo agregar si tiene contenido válido
        if case_data['Nombre'] or case_data['Secuencia']:
            parsed_cases.append(case_data)
    
    return parsed_cases

def parse_text_format(content):
    """
    Parsea formato de texto corrido (como tu pregunta original)
    """
    # Dividir por casos usando números al inicio
    cases = re.split(r'\n(?=\d+\s+Funcional)', content.strip())
    
    parsed_cases = []
    
    for case in cases:
        if not case.strip():
            continue
        
        case_data = parse_single_text_case(case.strip())
        if case_data:
            parsed_cases.append(case_data)
    
    return parsed_cases

def parse_single_text_case(text):
    """
    Parsea un caso individual de formato texto corrido
    """
    try:
        # Extraer nombre (hasta el primer "1.")
        nombre_match = re.match(r'^(.*?)\s*(?=1\.)', text)
        if nombre_match:
            nombre = nombre_match.group(1).strip()
        else:
            # Fallback: primeras 10 palabras
            words = text.split()
            nombre = ' '.join(words[:10]) if len(words) > 10 else text
        
        # Extraer secuencia (pasos numerados)
        secuencia_matches = re.findall(r'(\d+\.\s*[^1-9]*?)(?=\d+\.|Usuario|Se\s|Los\s|La\s|El\s|—|$)', text)
        secuencia = ' '.join(match.strip() for match in secuencia_matches)
        
        # Extraer datos adicionales
        datos_patterns = [
            r'Usuario\s+(?:no\s+)?suscrito(?:\s+activo)?(?:\s+autenticado)?',
            r'Artículo\s+[^.]*',
            r'—'  # Para casos Look & Feel
        ]
        
        datos_adicionales = ""
        for pattern in datos_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                datos_adicionales = match.group(0).strip()
                break
        
        # Extraer resultado esperado
        resultado_patterns = [
            r'(Se\s+(?:muestra|abre|muestran|inicia|permite|activa|visualiza|redirige|determina)[^.]*\.?)',
            r'(Los\s+íconos[^.]*\.?)',
            r'(La\s+(?:aplicación|redistribución)[^.]*\.?)',
            r'(El\s+(?:ícono|CTA|efecto|comportamiento|estatus|clic)[^.]*\.?)',
            r'(Imagen\s+destacada[^.]*\.?)'
        ]
        
        resultado_esperado = ""
        for pattern in resultado_patterns:
            match = re.search(pattern, text, re.IGNORECASE)
            if match:
                resultado_esperado = match.group(1).strip()
                break
        
        return {
            'Nombre': nombre,
            'Secuencia': secuencia,
            'Datos requeridos / Datos Adicionales': datos_adicionales,
            'Resultado Esperado': resultado_esperado
        }
    
    except Exception as e:
        print(f"Error parseando caso: {e}")
        return None

def create_formatted_excel(data, output_path):
    """
    Crea un archivo Excel con formato específico usando solo openpyxl
    """
    # Crear workbook y worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Casos de Prueba"
    
    # Definir estilos
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    alignment_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    alignment_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
    
    # Bordes
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Escribir headers
    headers = ['Nombre', 'Secuencia', 'Datos requeridos / Datos Adicionales', 'Resultado Esperado']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment_center
        cell.border = thin_border
    
    # Escribir datos
    for row_idx, case in enumerate(data, 2):
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=case.get(header, ''))
            cell.alignment = alignment_left
            cell.border = thin_border
    
    # Ajustar anchos de columna
    column_widths = {
        'A': 25,  # Nombre
        'B': 40,  # Secuencia  
        'C': 20,  # Datos requeridos
        'D': 35   # Resultado Esperado
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Ajustar altura de filas para contenido
    for row in range(2, len(data) + 2):
        ws.row_dimensions[row].height = 60
    
    # Guardar archivo
    wb.save(output_path)
    print(f"Excel generado exitosamente: {output_path}")

def main():
    """
    Función principal
    """
    # Configurar rutas
    input_directory = "./input"  # Directorio donde están los archivos txt
    output_directory = "./output"  # Directorio donde se guardará el Excel
    
    # Crear directorios si no existen
    os.makedirs(input_directory, exist_ok=True)
    os.makedirs(output_directory, exist_ok=True)
    
    # Buscar archivos .txt en el directorio input
    txt_files = [f for f in os.listdir(input_directory) if f.endswith('.txt')]
    
    if not txt_files:
        print(f"No se encontraron archivos .txt en {input_directory}")
        print("Por favor, coloca tu archivo de casos de prueba en formato .txt en esa carpeta")
        return
    
    # Procesar cada archivo encontrado
    for txt_file in txt_files:
        print(f"\nProcesando archivo: {txt_file}")
        
        input_path = os.path.join(input_directory, txt_file)
        timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = txt_file.replace('.txt', f'_casos_prueba_{timestamp}.xlsx')
        output_path = os.path.join(output_directory, output_file)
        
        # Parsear casos de prueba
        test_cases = parse_test_cases_from_file(input_path)
        
        if not test_cases:
            print(f"No se pudieron extraer casos de prueba de {txt_file}")
            continue
        
        print(f"Se encontraron {len(test_cases)} casos de prueba")
        
        # Generar Excel
        create_formatted_excel(test_cases, output_path)
        
        # Mostrar preview de los casos encontrados
        print("\nPreview de casos extraídos:")
        for i, case in enumerate(test_cases[:3], 1):  # Mostrar solo los primeros 3
            print(f"\n--- Caso {i} ---")
            print(f"Nombre: {case['Nombre'][:50]}...")
            print(f"Secuencia: {case['Secuencia'][:50]}...")
            print(f"Resultado: {case['Resultado Esperado'][:50]}...")

if __name__ == "__main__":
    main()